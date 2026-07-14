from .models import CD
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db.models import Sum
from django.contrib.auth import logout
from django.db.models import Q 
from django.http import HttpResponse
import openpyxl
import subprocess
from django.shortcuts import get_object_or_404
from django.shortcuts import get_object_or_404
from django.db.models import Count


import os

import openpyxl
from .models import CD


def logout_view(request):
    logout(request)
    return redirect("login")

@ensure_csrf_cookie
def login_view(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None:
            login(request, user)
            return redirect("dashboard")

        else:
            messages.error(request, "Invalid username or password")

    return render(request, "index.html")


from .models import CD

def dashboard(request):

    total_cds = CD.objects.count()

    total_categories = CD.objects.values("category").distinct().count()

    recent_cds = CD.objects.order_by("-created_at")[:5]

    # Total files
    total_files = sum(cd.number_of_files for cd in CD.objects.all())

    # Total storage (in MB)
    total_storage_mb = 0

    for cd in CD.objects.all():

        size = cd.size.strip().upper()

        try:
            if "GB" in size:
                total_storage_mb += float(size.replace("GB", "").strip()) * 1024

            elif "MB" in size:
                total_storage_mb += float(size.replace("MB", "").strip())

        except:
            pass

    # Convert to GB if needed
    if total_storage_mb >= 1024:
        total_storage = f"{total_storage_mb / 1024:.2f} GB"
    else:
        total_storage = f"{total_storage_mb:.2f} MB"

    context = {
        "total_cds": total_cds,
        "total_categories": total_categories,
        "recent_cds": recent_cds,
        "total_files": total_files,
        "total_storage": total_storage,
    }

    return render(request, "pages/dashboard.html", context)
def add_cd(request):

    if request.method == "POST":

        cd = CD(
            cd_label=request.POST.get("cd_label"),
            folder_name=request.POST.get("folder_name"),
            folder_path=request.POST.get("folder_path"),
            category=request.POST.get("category"),
            year=request.POST.get("year"),
            number_of_files=request.POST.get("number_of_files"),
            size=request.POST.get("size"),
            condition=request.POST.get("condition"),
            remarks=request.POST.get("remarks")
        )

        cd.save()

        return redirect("view_cd")

    return render(request, "pages/add_cd.html")


from .models import CD
from django.db.models import Q

def view_cd(request):

    search = request.GET.get("search")

    cds = CD.objects.all()

    if search:

        cds = cds.filter(
            Q(cd_label__icontains=search) |
            Q(folder_name__icontains=search) |
            Q(category__icontains=search) |
            Q(year__icontains=search)
        )

    return render(
        request,
        "pages/view_cd.html",
        {
            "cds": cds,
        }
    )

def edit_cd(request, id):

    cd = CD.objects.get(id=id)

    if request.method == "POST":

        cd.cd_label = request.POST.get("cd_label")
        cd.folder_name = request.POST.get("folder_name")
        cd.folder_path = request.POST.get("folder_path")
        cd.category = request.POST.get("category")
        cd.year = request.POST.get("year")
        cd.number_of_files = request.POST.get("number_of_files")
        cd.size = request.POST.get("size")
        cd.condition = request.POST.get("condition")
        cd.remarks = request.POST.get("remarks")

        cd.save()

        return redirect("view_cd")

    return render(
        request,
        "pages/edit_cd.html",
        {"cd": cd}
    )

def delete_cd(request, id):
    cd = CD.objects.get(id=id)
    cd.delete()

    messages.success(request, "CD deleted successfully!")

    return redirect("view_cd")

import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import CD


def import_excel(request):

    if request.method == "POST":

        excel_file = request.FILES["excel_file"]

        workbook = openpyxl.load_workbook(excel_file)

        sheet = workbook.active


        imported = 0
        updated = 0

        for row in sheet.iter_rows(min_row=5, values_only=True):

            # Skip empty rows
            if not row or not row[1]:
                continue

            # Skip summary rows
            if str(row[1]).startswith("Total") \
                    or str(row[1]).startswith("Good condition") \
                    or str(row[1]).startswith("Scratched") \
                    or str(row[1]).startswith("Unreadable"):
                continue

            folder_name = row[2] or ""



            obj, created = CD.objects.update_or_create(
                       cd_label=row[1],
                       defaults={
                              "folder_name": row[2] or "",
                              "folder_path": row[3] or "",
                              "category": row[4] or "",
                              "year": int(row[6]) if row[6] else 0,
                              "number_of_files": int(row[7]) if row[7] else 0,
                              "size": str(row[8]) if row[8] else "",
                              "condition": row[9] or "Good",
                              "remarks": row[11] or "",
                        }
                    )

            if created:
                imported += 1
            else:
                updated += 1

        messages.success(
            request,
            f"Import Completed! {imported} new CDs added, {updated} CDs updated."
        )

        return redirect("view_cd")

    return render(request, "pages/import_excel.html")

def export_excel(request):

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "CD Archive"

    # Header row
    sheet.append([
        "CD Label",
        "Folder Name",
        "Folder Path",
        "Category",
        "Year",
        "Number of Files",
        "Size",
        "Condition",
        "Remarks",
    ])

    # Data rows
    cds = CD.objects.all()

    for cd in cds:
        sheet.append([
            cd.cd_label,
            cd.folder_name,
            cd.folder_path,
            cd.category,
            cd.year,
            cd.number_of_files,
            cd.size,
            cd.condition,
            cd.remarks,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="CD_Archive.xlsx"'

    workbook.save(response)

    return response

import os
import subprocess
from django.shortcuts import get_object_or_404
from django.contrib import messages

def open_folder(request, id):

    cd = get_object_or_404(CD, id=id)

    if os.path.exists(cd.folder_path):

        subprocess.Popen(f'explorer "{cd.folder_path}"')

    else:

        messages.error(request, "Folder not found!")

    return redirect("view_cd")

    
def view_details(request, id):
    cd = get_object_or_404(CD, id=id)
    return render(request, "pages/view_details.html", {"cd": cd})



def categories(request):
    categories = (
        CD.objects
        .values("category")
        .annotate(total=Count("id"))
        .order_by("category")
    )

    return render(
        request,
        "pages/categories.html",
        {"categories": categories}
    )